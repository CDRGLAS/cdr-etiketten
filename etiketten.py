"""
CDR Glas AG – Handelsetiketten drucken
Liest AUF.xlsx ein, sucht nach Auftrag/Position, befüllt die Excel-Vorlage und öffnet sie.
"""
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from copy import copy
import openpyxl
from openpyxl.drawing.image import Image as XlImage

# ===== Paths =====
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'Etikette_Vorlage.xlsx')
LOGO_PATH = os.path.join(SCRIPT_DIR, 'CDR_Logo_pos_RGB-01.jpg')
OUTPUT_DE = os.path.join(SCRIPT_DIR, 'Etikette_Druck_DE.xlsx')
OUTPUT_FR = os.path.join(SCRIPT_DIR, 'Etikette_Druck_FR.xlsx')
PRINTER_NAME = 'CAB-EOS5/200 auf Ne04:'


def load_auf(path):
    """Load AUF.xlsx and return list of dicts keyed by (doknr, pos)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[1]:
            continue
        doknr = str(row[1]).strip()
        pos = str(row[2]).strip()
        komm_raw = str(row[3] or '')
        komm_parts = komm_raw.split('\n', 1)
        kommission = komm_parts[0].strip()
        kunden_pos = komm_parts[1].strip() if len(komm_parts) > 1 else ''
        produkt_raw = str(row[4] or '')
        produkt = produkt_raw.split('\n', 1)[0].strip()
        termin = row[13]
        if isinstance(termin, str):
            try:
                termin = datetime.strptime(termin, '%Y-%m-%d')
            except ValueError:
                termin = None

        data[(doknr, pos)] = {
            'doknr': doknr,
            'pos': pos,
            'kommission': kommission,
            'kunden_pos': kunden_pos,
            'produkt': produkt,
            'qm': float(row[5] or 0),
            'lfdm': float(row[6] or 0),
            'gewicht': float(row[7] or 0),
            'kunde': str(row[11] or '').strip(),
            'termin': termin,
            'menge': int(row[17] or 0),
            'breite': int(row[18] or 0),
            'hoehe': int(row[19] or 0),
        }
    wb.close()
    return data


def build_barcode_text(doknr, pos):
    """Build Code-39 barcode text: *AuftragNr + Pos(5-digit)*"""
    return '*' + str(doknr) + str(pos).zfill(5) + '*'


def fill_template(record, lang='de'):
    """Fill the Excel template with data and return the workbook."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_de = wb['Glas-Etikette']

    barcode = build_barcode_text(record['doknr'], record['pos'])

    # Barcode cells (replace formulas with computed value)
    ws_de['E1'] = barcode
    ws_de['K1'] = barcode

    # AU/Pos
    ws_de['G4'] = int(record['doknr'])
    ws_de['H4'] = int(record['pos'])

    # I1 = Auftrag-Nr (was formula =G4)
    ws_de['I1'] = int(record['doknr'])

    # I9 = Position (was formula =H4)
    ws_de['I9'] = int(record['pos'])

    # Termin
    ws_de['G5'] = record['termin']

    # Kunde
    ws_de['A8'] = record['kunde']

    # Kommission
    ws_de['A11'] = record['kommission']

    # Kunden-Pos.
    ws_de['A14'] = record['kunden_pos']

    # Produkt
    ws_de['A17'] = record['produkt']

    # Menge
    ws_de['B19'] = record['menge']

    # Breite, Höhe
    ws_de['E19'] = record['breite']
    ws_de['H19'] = record['hoehe']

    # Gewicht (ganzzahlig)
    ws_de['H20'] = round(record['gewicht'])

    # K15 - ID number (leave as-is or clear for Handelsware)
    ws_de['K15'] = ''

    # Fläche und Umfang: Formeln bleiben, aber auch direkte Werte setzen falls nötig
    # B20 hat Formel =(E19/1000)*(H19/1000) - lassen wir
    # E20 hat Formel =(E19/1000)*2+(H19/1000)*2 - lassen wir

    # FR sheet: fix formula references (replace =ID!... with values)
    ws_fr = wb['Glas-Etikette FR']
    ws_fr['E1'] = barcode
    ws_fr['K1'] = barcode
    ws_fr['K15'] = ''

    # Add CDR Logo
    if os.path.exists(LOGO_PATH):
        for ws in [ws_de, ws_fr]:
            logo = XlImage(LOGO_PATH)
            logo.width = 120
            logo.height = 60
            ws.add_image(logo, 'A1')

    return wb


def drucken(record, lang='de'):
    """Fill template, save, and print directly via Excel COM."""
    wb = fill_template(record, lang)

    output_path = OUTPUT_FR if lang == 'fr' else OUTPUT_DE
    sheet_name = 'Glas-Etikette FR' if lang == 'fr' else 'Glas-Etikette'

    if lang == 'fr':
        wb.active = wb.sheetnames.index('Glas-Etikette FR')
    else:
        wb.active = wb.sheetnames.index('Glas-Etikette')

    wb.save(output_path)
    wb.close()

    # Drucken wie VBA: PrintOut ActivePrinter:="CAB-EOS5/200"
    import win32com.client

    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        xwb = excel.Workbooks.Open(os.path.abspath(output_path))
        xws = xwb.Worksheets(sheet_name)
        copies = int(record.get('menge', 1)) or 1
        xws.PrintOut(Copies=copies, ActivePrinter=PRINTER_NAME)
        xwb.Close(SaveChanges=False)
    finally:
        excel.DisplayAlerts = True
        excel.Quit()


# ===== GUI =====
class EtikettenApp:
    def __init__(self, root):
        self.root = root
        self.root.title('CDR Glas AG – Handelsetiketten')
        self.root.resizable(True, False)
        self.auf_data = {}

        # Style
        style = ttk.Style()
        style.configure('Header.TLabel', font=('Segoe UI', 11, 'bold'))
        style.configure('Status.TLabel', font=('Segoe UI', 9))

        main = ttk.Frame(root, padding=16)
        main.grid(row=0, column=0, sticky='nsew')

        # --- File loading ---
        ttk.Label(main, text='Handelsetiketten drucken', style='Header.TLabel').grid(
            row=0, column=0, columnspan=4, pady=(0, 12), sticky='w')

        ttk.Button(main, text='AUF.xlsx laden...', command=self.load_file).grid(
            row=1, column=0, columnspan=2, sticky='w')
        self.file_status = ttk.Label(main, text='Keine Datei geladen', style='Status.TLabel')
        self.file_status.grid(row=1, column=2, columnspan=2, sticky='w', padx=(8, 0))

        ttk.Separator(main, orient='horizontal').grid(
            row=2, column=0, columnspan=4, sticky='ew', pady=10)

        # --- Lookup ---
        ttk.Label(main, text='Auftrag-Nr.').grid(row=3, column=0, sticky='w')
        self.var_auftrag = tk.StringVar()
        e_auf = ttk.Entry(main, textvariable=self.var_auftrag, width=14)
        e_auf.grid(row=3, column=1, sticky='w', padx=(0, 8))
        self.var_auftrag.trace_add('write', lambda *a: self.auto_suchen())

        ttk.Label(main, text='Position').grid(row=3, column=2, sticky='w')
        self.var_pos = tk.StringVar()
        e_pos = ttk.Entry(main, textvariable=self.var_pos, width=8)
        e_pos.grid(row=3, column=3, sticky='w')
        self.var_pos.trace_add('write', lambda *a: self.auto_suchen())

        self.lookup_msg = ttk.Label(main, text='', style='Status.TLabel')
        self.lookup_msg.grid(row=4, column=0, columnspan=4, sticky='w', pady=(4, 0))

        ttk.Separator(main, orient='horizontal').grid(
            row=5, column=0, columnspan=4, sticky='ew', pady=10)

        # --- Data fields ---
        fields = [
            ('Termin', 'termin', 6),
            ('Menge', 'menge', 7),
            ('Kunde', 'kunde', 8),
            ('Kommission', 'kommission', 9),
            ('Kunden-Pos.', 'kunden_pos', 10),
            ('Produkt', 'produkt', 11),
            ('Breite [mm]', 'breite', 12),
            ('Höhe [mm]', 'hoehe', 13),
            ('Gewicht [kg]', 'gewicht', 14),
        ]

        self.vars = {}
        for label_text, key, row_num in fields:
            ttk.Label(main, text=label_text).grid(row=row_num, column=0, sticky='w', pady=2)
            var = tk.StringVar()
            self.vars[key] = var
            span = 3 if key in ('kunde', 'kommission', 'kunden_pos', 'produkt') else 1
            col_start = 1
            entry = ttk.Entry(main, textvariable=var, width=70 if span > 1 else 14)
            entry.grid(row=row_num, column=col_start, columnspan=span, sticky='we', pady=2)

        ttk.Separator(main, orient='horizontal').grid(
            row=15, column=0, columnspan=4, sticky='ew', pady=10)

        # --- Buttons ---
        btn_frame = ttk.Frame(main)
        btn_frame.grid(row=16, column=0, columnspan=4, sticky='w')

        ttk.Button(btn_frame, text='Drucken DE', command=lambda: self.drucken('de')).pack(
            side='left', padx=(0, 8))
        ttk.Button(btn_frame, text='Drucken FR', command=lambda: self.drucken('fr')).pack(
            side='left', padx=(0, 8))
        ttk.Button(btn_frame, text='Leeren', command=self.leeren).pack(side='left')

        # Try to auto-load AUF.xlsx from same directory
        default_auf = os.path.join(SCRIPT_DIR, 'AUF.xlsx')
        if os.path.exists(default_auf):
            self._load(default_auf)

    def load_file(self):
        path = filedialog.askopenfilename(
            title='AUF.xlsx auswählen',
            filetypes=[('Excel-Dateien', '*.xlsx *.xls')],
            initialdir=SCRIPT_DIR
        )
        if path:
            self._load(path)

    def _load(self, path):
        try:
            self.auf_data = load_auf(path)
            name = os.path.basename(path)
            self.file_status.config(text=f'{name} – {len(self.auf_data)} Aufträge geladen')
        except Exception as e:
            messagebox.showerror('Fehler', f'Datei konnte nicht geladen werden:\n{e}')

    def auto_suchen(self):
        """Automatisch suchen, sobald Auftrag-Nr. und Position eingegeben sind."""
        nr = self.var_auftrag.get().strip()
        pos = self.var_pos.get().strip()
        if nr and pos and self.auf_data:
            self.suchen()

    def suchen(self):
        nr = self.var_auftrag.get().strip()
        pos = self.var_pos.get().strip()

        if not nr:
            self.lookup_msg.config(text='Bitte Auftrag-Nr. eingeben.')
            return

        # Try exact match first, then try without position
        key = (nr, pos)
        record = self.auf_data.get(key)

        if not record and not pos:
            # Find first match for this Auftrag-Nr
            for k, v in self.auf_data.items():
                if k[0] == nr:
                    record = v
                    self.var_pos.set(v['pos'])
                    break

        if not record:
            self.lookup_msg.config(text=f'Auftrag {nr}/{pos} nicht gefunden.')
            return

        # Fill fields
        if record['termin']:
            self.vars['termin'].set(record['termin'].strftime('%d.%m.%Y'))
        self.vars['menge'].set(str(record['menge']))
        self.vars['kunde'].set(record['kunde'])
        self.vars['kommission'].set(record['kommission'])
        self.vars['kunden_pos'].set(record['kunden_pos'])
        self.vars['produkt'].set(record['produkt'])
        self.vars['breite'].set(str(record['breite']))
        self.vars['hoehe'].set(str(record['hoehe']))
        self.vars['gewicht'].set(str(round(record['gewicht'])))

        self.lookup_msg.config(text=f'Geladen: {record["doknr"]}/{record["pos"]} – {record["kunde"]}')

    def get_record(self):
        """Build a record dict from the form fields."""
        termin_str = self.vars['termin'].get().strip()
        termin = None
        if termin_str:
            for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y'):
                try:
                    termin = datetime.strptime(termin_str, fmt)
                    break
                except ValueError:
                    continue

        breite = int(self.vars['breite'].get() or 0)
        hoehe = int(self.vars['hoehe'].get() or 0)

        return {
            'doknr': self.var_auftrag.get().strip(),
            'pos': self.var_pos.get().strip() or '1',
            'termin': termin,
            'menge': int(self.vars['menge'].get() or 1),
            'kunde': self.vars['kunde'].get().strip(),
            'kommission': self.vars['kommission'].get().strip(),
            'kunden_pos': self.vars['kunden_pos'].get().strip(),
            'produkt': self.vars['produkt'].get().strip(),
            'breite': breite,
            'hoehe': hoehe,
            'gewicht': float(self.vars['gewicht'].get() or 0),
            'qm': breite * hoehe / 1000000 if breite and hoehe else 0,
            'lfdm': 2 * (breite + hoehe) / 1000 if breite and hoehe else 0,
        }

    def drucken(self, lang):
        record = self.get_record()
        if not record['doknr']:
            messagebox.showwarning('Fehler', 'Bitte Auftrag-Nr. eingeben.')
            return
        try:
            drucken(record, lang)
        except Exception as e:
            messagebox.showerror('Fehler', f'Fehler beim Erstellen der Etikette:\n{e}')

    def leeren(self):
        self.var_auftrag.set('')
        self.var_pos.set('')
        for var in self.vars.values():
            var.set('')
        self.lookup_msg.config(text='')


if __name__ == '__main__':
    root = tk.Tk()
    app = EtikettenApp(root)
    root.mainloop()
