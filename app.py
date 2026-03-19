"""
CDR Glas AG – Handelsetiketten drucken (Web-App)
Flask-basierte Web-Applikation.
"""
import os
from datetime import datetime
from copy import copy
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from flask import Flask, render_template_string, request, jsonify

# ===== Paths =====
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'Etikette_Vorlage.xlsx')
LOGO_PATH = os.path.join(SCRIPT_DIR, 'CDR_Logo_pos_RGB-01.jpg')
PRINTER_NAME = 'CAB-EOS5/200 auf Ne04:'

# ===== Data =====
auf_data = {}


def load_auf(path):
    """Load AUF.xlsx and return dict keyed by (doknr, pos)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        doknr = str(row[1]).strip()
        pos = str(row[2]).strip()
        komm_raw = str(row[3] or '')
        komm_parts = komm_raw.split('\n', 1)
        kommission = komm_parts[0].strip()
        kunden_pos = komm_parts[1].strip() if len(komm_parts) > 1 else ''
        produkt_raw = str(row[4] or '')
        produkt = produkt_raw.split('\n', 1)[0].strip()
        termin = row[13]
        if isinstance(termin, str):
            try:
                termin = datetime.strptime(termin, '%Y-%m-%d')
            except ValueError:
                termin = None

        data[(doknr, pos)] = {
            'doknr': doknr,
            'pos': pos,
            'kommission': kommission,
            'kunden_pos': kunden_pos,
            'produkt': produkt,
            'qm': float(row[5] or 0),
            'lfdm': float(row[6] or 0),
            'gewicht': float(row[7] or 0),
            'kunde': str(row[11] or '').strip(),
            'termin': termin,
            'menge': int(row[17] or 0),
            'breite': int(row[18] or 0),
            'hoehe': int(row[19] or 0),
        }
    wb.close()
    return data


def build_barcode_text(doknr, pos):
    return '*' + str(doknr) + str(pos).zfill(5) + '*'


def fill_template(record):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_de = wb['Glas-Etikette']
    barcode = build_barcode_text(record['doknr'], record['pos'])

    ws_de['E1'] = barcode
    ws_de['K1'] = barcode
    ws_de['G4'] = int(record['doknr'])
    ws_de['H4'] = int(record['pos'])
    ws_de['I1'] = int(record['doknr'])
    ws_de['I9'] = int(record['pos'])
    ws_de['G5'] = record['termin']
    ws_de['A8'] = record['kunde']
    ws_de['A11'] = record['kommission']
    ws_de['A14'] = record['kunden_pos']
    ws_de['A17'] = record['produkt']
    ws_de['B19'] = record['menge']
    ws_de['E19'] = record['breite']
    ws_de['H19'] = record['hoehe']
    ws_de['H20'] = round(record['gewicht'])
    ws_de['K15'] = ''

    ws_fr = wb['Glas-Etikette FR']
    ws_fr['E1'] = barcode
    ws_fr['K1'] = barcode
    ws_fr['K15'] = ''

    if os.path.exists(LOGO_PATH):
        for ws in [ws_de, ws_fr]:
            logo = XlImage(LOGO_PATH)
            logo.width = 120
            logo.height = 60
            ws.add_image(logo, 'A1')

    return wb


def drucken(record, lang='de'):
    """Vorlage direkt in Excel öffnen, Zellen befüllen, drucken, ohne Speichern schliessen."""
    import win32com.client
    sheet_name = 'Glas-Etikette FR' if lang == 'fr' else 'Glas-Etikette'
    barcode = build_barcode_text(record['doknr'], record['pos'])

    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        xwb = excel.Workbooks.Open(os.path.abspath(TEMPLATE_PATH))
        ws = xwb.Worksheets('Glas-Etikette')

        # Zellen befüllen
        ws.Range('E1').Value = barcode
        ws.Range('K1').Value = barcode
        ws.Range('G4').Value = int(record['doknr'])
        ws.Range('H4').Value = int(record['pos'])
        ws.Range('I1').Value = int(record['doknr'])
        ws.Range('I9').Value = int(record['pos'])
        ws.Range('G5').Value = record['termin']
        ws.Range('A8').Value = record['kunde']
        ws.Range('A11').Value = record['kommission']
        ws.Range('A14').Value = record['kunden_pos']
        ws.Range('A17').Value = record['produkt']
        ws.Range('B19').Value = record['menge']
        ws.Range('E19').Value = record['breite']
        ws.Range('H19').Value = record['hoehe']
        ws.Range('H20').Value = round(record['gewicht'])
        ws.Range('K15').Value = ''

        # FR barcode cells
        ws_fr = xwb.Worksheets('Glas-Etikette FR')
        ws_fr.Range('E1').Value = barcode
        ws_fr.Range('K1').Value = barcode
        ws_fr.Range('K15').Value = ''

        # Drucken
        xws = xwb.Worksheets(sheet_name)
        copies = int(record.get('menge', 1)) or 1
        xws.PrintOut(Copies=copies, ActivePrinter=PRINTER_NAME)

        # Schliessen ohne Speichern
        xwb.Close(SaveChanges=False)
    finally:
        excel.DisplayAlerts = True
        excel.Quit()


# ===== Flask App =====
app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CDR Glas AG – Handelsetiketten</title>
<style>
  :root {
    --cdr-blue: #003366;
    --cdr-light: #f4f6f8;
    --cdr-border: #d0d5dd;
    --cdr-green: #1a7a3a;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Segoe UI', Tahoma, sans-serif;
    background: var(--cdr-light);
    color: #222;
  }
  .app-header {
    background: var(--cdr-blue);
    color: white;
    padding: 14px 24px;
    display: flex;
    align-items: center;
    gap: 16px;
  }
  .app-header h1 { font-size: 18px; font-weight: 600; }
  .app-header span { font-size: 13px; opacity: 0.7; }

  .container {
    max-width: 800px;
    margin: 24px auto;
    background: white;
    border-radius: 8px;
    border: 1px solid var(--cdr-border);
    padding: 24px;
  }

  h2 {
    font-size: 15px;
    color: var(--cdr-blue);
    margin-bottom: 16px;
    border-bottom: 2px solid var(--cdr-blue);
    padding-bottom: 6px;
  }

  .lookup-row {
    display: flex;
    gap: 12px;
    align-items: flex-end;
    margin-bottom: 8px;
  }
  .lookup-msg {
    font-size: 13px;
    min-height: 20px;
    margin-bottom: 16px;
    padding: 4px 0;
  }
  .lookup-msg.success { color: var(--cdr-green); font-weight: 600; }
  .lookup-msg.error { color: #c0392b; }

  .form-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px 20px;
  }
  .form-grid .full { grid-column: 1 / -1; }

  .form-group label {
    display: block;
    font-size: 11px;
    font-weight: 600;
    color: #555;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 3px;
  }
  .form-group input {
    width: 100%;
    padding: 8px 10px;
    border: 1px solid var(--cdr-border);
    border-radius: 4px;
    font-size: 14px;
    font-family: inherit;
  }
  .form-group input:focus {
    outline: none;
    border-color: var(--cdr-blue);
    box-shadow: 0 0 0 2px rgba(0,51,102,0.15);
  }
  .form-group .computed {
    background: #f0f4f8;
    color: #555;
  }

  .btn-row {
    margin-top: 20px;
    display: flex;
    gap: 10px;
  }
  .btn {
    padding: 10px 24px;
    border: none;
    border-radius: 4px;
    font-size: 14px;
    font-weight: 600;
    cursor: pointer;
    transition: background 0.15s;
  }
  .btn-primary { background: var(--cdr-blue); color: white; }
  .btn-primary:hover { background: #004488; }
  .btn-primary:disabled { opacity: 0.5; cursor: wait; }
  .btn-secondary { background: #e8ecf0; color: #333; }
  .btn-secondary:hover { background: #d0d5dd; }

  .status-bar {
    margin-top: 12px;
    padding: 8px 12px;
    border-radius: 4px;
    font-size: 13px;
    display: none;
  }
  .status-bar.success { display: block; background: #e8f5e9; color: #2e7d32; }
  .status-bar.error { display: block; background: #ffebee; color: #c62828; }
  .status-bar.printing { display: block; background: #e3f2fd; color: #1565c0; }

  .file-info {
    font-size: 13px;
    color: var(--cdr-green);
    font-weight: 600;
    margin-bottom: 16px;
  }
</style>
</head>
<body>

<div class="app-header">
  <h1>Handelsetiketten drucken</h1>
  <span>CDR Glas AG</span>
</div>

<div class="container">
  <div class="file-info" id="fileInfo">{{ file_info }}</div>

  <div class="lookup-row">
    <div class="form-group">
      <label>Auftrag-Nr.</label>
      <input type="text" id="auftragNr" placeholder="z.B. 142830">
    </div>
    <div class="form-group">
      <label>Position</label>
      <input type="text" id="position" placeholder="z.B. 1">
    </div>
  </div>
  <div class="lookup-msg" id="lookupMsg"></div>

  <h2>Etikettendaten</h2>
  <div class="form-grid">
    <div class="form-group">
      <label>Termin</label>
      <input type="date" id="termin">
    </div>
    <div class="form-group">
      <label>Menge [Stk]</label>
      <input type="number" id="menge" value="1" min="1">
    </div>
    <div class="form-group full">
      <label>Kunde</label>
      <input type="text" id="kunde">
    </div>
    <div class="form-group full">
      <label>Kommission</label>
      <input type="text" id="kommission">
    </div>
    <div class="form-group full">
      <label>Kunden-Pos.</label>
      <input type="text" id="kundenPos">
    </div>
    <div class="form-group full">
      <label>Produkt</label>
      <input type="text" id="produkt">
    </div>
    <div class="form-group">
      <label>Breite [mm]</label>
      <input type="number" id="breite" min="0">
    </div>
    <div class="form-group">
      <label>Höhe [mm]</label>
      <input type="number" id="hoehe" min="0">
    </div>
    <div class="form-group">
      <label>Fläche [m²]</label>
      <input type="text" id="flaeche" class="computed" readonly>
    </div>
    <div class="form-group">
      <label>Umfang [lfm]</label>
      <input type="text" id="umfang" class="computed" readonly>
    </div>
    <div class="form-group">
      <label>Gewicht [kg]</label>
      <input type="number" id="gewicht" min="0" step="1">
    </div>
  </div>

  <div class="btn-row">
    <button class="btn btn-primary" onclick="druckenPos('de')">DE Position</button>
    <button class="btn btn-primary" onclick="druckenAlle('de')">DE alle Pos.</button>
    <button class="btn btn-primary" onclick="druckenPos('fr')">FR Position</button>
    <button class="btn btn-primary" onclick="druckenAlle('fr')">FR alle Pos.</button>
    <button class="btn btn-secondary" onclick="leeren()">Leeren</button>
  </div>

  <div class="status-bar" id="statusBar"></div>
</div>

<script>
  const fields = ['termin','menge','kunde','kommission','kundenPos','produkt','breite','hoehe','gewicht','flaeche','umfang'];

  // Auto-search
  let searchTimer = null;
  document.getElementById('auftragNr').addEventListener('input', scheduleSearch);
  document.getElementById('position').addEventListener('input', scheduleSearch);

  function scheduleSearch() {
    clearTimeout(searchTimer);
    searchTimer = setTimeout(autoSearch, 300);
  }

  function autoSearch() {
    const nr = document.getElementById('auftragNr').value.trim();
    const pos = document.getElementById('position').value.trim();
    if (!nr || !pos) return;

    fetch('/api/suchen?' + new URLSearchParams({nr, pos}))
      .then(r => r.json())
      .then(data => {
        const msg = document.getElementById('lookupMsg');
        if (data.error) {
          msg.className = 'lookup-msg error';
          msg.textContent = data.error;
          return;
        }
        msg.className = 'lookup-msg success';
        msg.textContent = 'Geladen: ' + data.doknr + '/' + data.pos + ' – ' + data.kunde;

        document.getElementById('termin').value = data.termin || '';
        document.getElementById('menge').value = data.menge || 1;
        document.getElementById('kunde').value = data.kunde || '';
        document.getElementById('kommission').value = data.kommission || '';
        document.getElementById('kundenPos').value = data.kunden_pos || '';
        document.getElementById('produkt').value = data.produkt || '';
        document.getElementById('breite').value = data.breite || '';
        document.getElementById('hoehe').value = data.hoehe || '';
        document.getElementById('gewicht').value = data.gewicht || '';
        document.getElementById('flaeche').value = data.qm ? data.qm.toFixed(2) : '';
        document.getElementById('umfang').value = data.lfdm ? data.lfdm.toFixed(2) : '';
      });
  }

  function getFormData() {
    return {
      doknr: document.getElementById('auftragNr').value.trim(),
      pos: document.getElementById('position').value.trim() || '1',
      termin: document.getElementById('termin').value,
      menge: parseInt(document.getElementById('menge').value) || 1,
      kunde: document.getElementById('kunde').value.trim(),
      kommission: document.getElementById('kommission').value.trim(),
      kunden_pos: document.getElementById('kundenPos').value.trim(),
      produkt: document.getElementById('produkt').value.trim(),
      breite: parseInt(document.getElementById('breite').value) || 0,
      hoehe: parseInt(document.getElementById('hoehe').value) || 0,
      gewicht: parseFloat(document.getElementById('gewicht').value) || 0,
      flaeche: parseFloat(document.getElementById('flaeche').value) || 0,
      umfang: parseFloat(document.getElementById('umfang').value) || 0,
    };
  }

  function setButtons(disabled) {
    document.querySelectorAll('.btn-primary').forEach(b => b.disabled = disabled);
  }

  function druckenPos(lang) {
    const data = getFormData();
    if (!data.doknr) { showStatus('Bitte Auftrag-Nr. eingeben.', 'error'); return; }

    showStatus('Drucke Position ' + data.pos + ' ' + lang.toUpperCase() + '...', 'printing');
    setButtons(true);

    fetch('/api/drucken', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({record: data, lang: lang})
    })
    .then(r => r.json())
    .then(result => {
      if (result.ok) {
        showStatus(result.menge + 'x Etikette ' + lang.toUpperCase() + ' (Pos ' + data.pos + ') gedruckt', 'success');
      } else {
        showStatus('Fehler: ' + result.error, 'error');
      }
    })
    .catch(err => showStatus('Fehler: ' + err, 'error'))
    .finally(() => setButtons(false));
  }

  function druckenAlle(lang) {
    const nr = document.getElementById('auftragNr').value.trim();
    if (!nr) { showStatus('Bitte Auftrag-Nr. eingeben.', 'error'); return; }

    showStatus('Drucke alle Positionen von Auftrag ' + nr + ' ' + lang.toUpperCase() + '...', 'printing');
    setButtons(true);

    fetch('/api/drucken_alle', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({doknr: nr, lang: lang})
    })
    .then(r => r.json())
    .then(result => {
      if (result.ok) {
        showStatus(result.total + ' Etiketten ' + lang.toUpperCase() + ' gedruckt (' + result.positionen + ' Positionen)', 'success');
      } else {
        showStatus('Fehler: ' + result.error, 'error');
      }
    })
    .catch(err => showStatus('Fehler: ' + err, 'error'))
    .finally(() => setButtons(false));
  }

  function leeren() {
    document.getElementById('auftragNr').value = '';
    document.getElementById('position').value = '';
    fields.forEach(id => document.getElementById(id).value = '');
    document.getElementById('menge').value = '1';
    document.getElementById('lookupMsg').textContent = '';
    document.getElementById('statusBar').className = 'status-bar';
  }

  function showStatus(text, type) {
    const bar = document.getElementById('statusBar');
    bar.textContent = text;
    bar.className = 'status-bar ' + type;
  }
</script>
</body>
</html>"""


@app.route('/')
def index():
    info = f'AUF.xlsx – {len(auf_data)} Aufträge geladen (R:\\CDR-Glas\\Lagerverwaltung)' if auf_data else 'Keine Datei geladen'
    return render_template_string(HTML, file_info=info)


@app.route('/api/suchen')
def api_suchen():
    nr = request.args.get('nr', '').strip()
    pos = request.args.get('pos', '').strip()

    if not nr:
        return jsonify({'error': 'Bitte Auftrag-Nr. eingeben.'})

    record = auf_data.get((nr, pos))
    if not record and not pos:
        for k, v in auf_data.items():
            if k[0] == nr:
                record = v
                break

    if not record:
        return jsonify({'error': f'Auftrag {nr}/{pos} nicht gefunden.'})

    result = dict(record)
    if result['termin']:
        result['termin'] = result['termin'].strftime('%Y-%m-%d')
    else:
        result['termin'] = ''
    result['gewicht'] = round(result['gewicht'])
    return jsonify(result)


@app.route('/api/drucken', methods=['POST'])
def api_drucken():
    data = request.json
    record = data['record']
    lang = data.get('lang', 'de')

    # Parse termin
    if record.get('termin'):
        try:
            record['termin'] = datetime.strptime(record['termin'], '%Y-%m-%d')
        except ValueError:
            record['termin'] = None
    else:
        record['termin'] = None

    try:
        drucken(record, lang)
        return jsonify({
            'ok': True,
            'menge': record.get('menge', 1),
            'printer': PRINTER_NAME
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/drucken_alle', methods=['POST'])
def api_drucken_alle():
    data = request.json
    doknr = str(data['doknr']).strip()
    lang = data.get('lang', 'de')

    # Alle Positionen zu diesem Auftrag finden
    positionen = [v for k, v in auf_data.items() if k[0] == doknr]

    if not positionen:
        return jsonify({'ok': False, 'error': f'Keine Positionen für Auftrag {doknr} gefunden.'})

    total = 0
    try:
        for record in positionen:
            rec = dict(record)
            drucken(rec, lang)
            total += rec.get('menge', 1)
        return jsonify({
            'ok': True,
            'total': total,
            'positionen': len(positionen),
            'printer': PRINTER_NAME
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


if __name__ == '__main__':
    # AUF.xlsx vom Netzlaufwerk laden
    auf_path = r'R:\CDR-Glas\Lagerverwaltung\AUF.xlsx'
    if os.path.exists(auf_path):
        auf_data = load_auf(auf_path)
        print(f'AUF.xlsx geladen: {len(auf_data)} Aufträge von {auf_path}')
    else:
        print(f'WARNUNG: {auf_path} nicht gefunden!')

    print('Starte Web-App auf http://localhost:5000')
    app.run(host='0.0.0.0', port=5000, debug=False)
